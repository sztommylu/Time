"""
Excel处理模块
负责Excel文件的读写、格式化和样式设置
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.dimensions import ColumnDimension
from config import (
    COLUMN_WIDTHS, COLOR_RULES, LIGHT_RED, DARK_RED, BRIGHT_RED, 
    LIGHT_GREEN, DARK_GREEN
)

class ExcelHandler:
    """Excel文件处理器"""
    
    def __init__(self, file_path):
        """
        初始化Excel处理器
        
        Args:
            file_path: Excel文件路径
        """
        self.file_path = file_path
    
    def read_data_from_sheet(self, sheet_name):
        """
        从指定工作表读取数据
        
        Args:
            sheet_name: 工作表名称
            
        Returns:
            DataFrame: 读取的数据，失败时返回None
        """
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
            print(f"成功从工作表读取数据: {sheet_name}")
            return df
        except Exception as e:
            print(f"从工作表 {sheet_name} 读取数据错误: {e}")
            return None
    
    def save_data_to_sheet(self, data, sheet_name, mode='w'):
        """
        将数据保存到指定工作表
        
        Args:
            data: 要保存的DataFrame数据
            sheet_name: 目标工作表名称
            mode: 写入模式，'w'为覆盖，'a'为追加（默认覆盖）
        """
        try:
            if mode == 'a':
                # 追加模式
                with pd.ExcelWriter(self.file_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
                    data.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                # 覆盖模式（默认）
                with pd.ExcelWriter(self.file_path, engine="openpyxl") as writer:
                    data.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"数据已成功保存到工作表: {sheet_name}")
        except Exception as e:
            print(f"保存数据到工作表 {sheet_name} 错误: {e}")

    def append_data_to_sheet(self, data, target_sheet_name):
        """
        将数据追加到指定工作表
        
        Args:
            data: 要写入的DataFrame数据
            target_sheet_name: 目标工作表名称
        """
        try:
            with pd.ExcelWriter(
                self.file_path, 
                mode="a", 
                engine="openpyxl", 
                if_sheet_exists="replace"
            ) as writer:
                data.to_excel(writer, sheet_name=target_sheet_name, index=False)
            print(f"数据成功追加到工作表: {target_sheet_name}")
        except Exception as e:
            print(f"追加数据到工作表 {target_sheet_name} 错误: {e}")
    
    def apply_color_coding(self, sheet_name):
        """
        为工作表应用颜色编码
        
        Args:
            sheet_name: 工作表名称
        """
        try:
            wb = load_workbook(self.file_path)
            ws = wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    # 如果是I列则跳过
                    if cell.column_letter == 'I':
                        continue
                    # 检查单元格的值是否为数字
                    if isinstance(cell.value, (int, float)):
                        value = cell.value
                        # 根据数值范围设置颜色
                        for condition, color in COLOR_RULES:
                            if condition(value):
                                cell.fill = color
                                break
            
            wb.save(self.file_path)
            print(f"颜色编码已应用到工作表: {sheet_name}")
            
        except Exception as e:
            print(f"应用颜色编码错误: {e}")
    
    def add_hyperlinks(self, sheet_name):
        """
        为股票名称添加超链接
        
        Args:
            sheet_name: 工作表名称
        """
        try:
            wb = load_workbook(self.file_path)
            ws = wb[sheet_name]
            
            # 遍历所有行（从第二行开始，假设第一行是表头）
            for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
                stock_code_cell = row[0]  # 假设股票代码在第一列
                stock_name_cell = row[1]  # 假设股票名称在第二列
                
                if stock_code_cell.value:
                    stock_code = str(stock_code_cell.value).split('.')[1]
                    
                    # 创建超链接
                    hyperlink = f"https://stockpage.10jqka.com.cn/{stock_code}/"
                    stock_name_cell.hyperlink = hyperlink
                    stock_name_cell.font = Font(color="0000FF", underline="single")  # 蓝色字体，带下划线
                    stock_name_cell.style = "Hyperlink"  # 应用超链接样式
            
            wb.save(self.file_path)
            print(f"超链接已添加到{sheet_name} 工作表")
            
        except Exception as e:
            print(f"添加超链接错误: {e}")
    
    def set_column_widths(self, sheet_name):
        """
        按照固定列字母顺序设置Excel列宽
        
        Args:
            sheet_name: 工作表名称
        """
        try:
            wb = load_workbook(self.file_path)
            ws = wb[sheet_name]
            
            # 确保所有列维度都存在
            for col_letter in COLUMN_WIDTHS.keys():
                if col_letter not in ws.column_dimensions:
                    ws.column_dimensions[col_letter] = ColumnDimension(ws, col_letter)
            
            # 设置列宽
            for col_letter, width in COLUMN_WIDTHS.items():
                ws.column_dimensions[col_letter].width = width
                print(f"设置列 {col_letter} 宽度为 {width}")
            
            wb.save(self.file_path)
            print(f"列宽已按照固定顺序设置到{sheet_name} 工作表")
            
        except Exception as e:
            print(f"设置列宽错误: {e}")
    
    def get_column_widths(self, sheet_name):
        """
        获取工作表的列宽信息
        
        Args:
            sheet_name: 工作表名称
            
        Returns:
            dict: 列名和列宽的映射
        """
        try:
            wb = load_workbook(self.file_path)
            ws = wb[sheet_name]
            
            column_widths = {}
            for column in ws.columns:
                column_letter = column[0].column_letter
                column_width = ws.column_dimensions[column_letter].width
                column_name = column[0].value 
                column_widths[column_name] = column_width
            
            return column_widths
            
        except Exception as e:
            print(f"获取列宽错误: {e}")
            return {}
    
    def format_worksheet(self, sheet_name):
        """
        对工作表进行完整的格式化处理
        包括颜色编码、超链接和列宽设置
        
        Args:
            sheet_name: 工作表名称
        """
        print(f"开始格式化工作表: {sheet_name}")
        self.add_hyperlinks(sheet_name)
        self.apply_color_coding(sheet_name)
        self.set_column_widths(sheet_name)
        print(f"工作表格式化完成: {sheet_name}")
    
    def read_filtered_data(self, sheet_name, filter_conditions=None):
        """
        读取并过滤数据
        
        Args:
            sheet_name: 工作表名称
            filter_conditions: 过滤条件函数
            
        Returns:
            list: 过滤后的数据列表
        """
        try:
            # skiprows=1 跳过第一行
            df = pd.read_excel(self.file_path, header=None, skiprows=1, sheet_name=sheet_name)
            
            if filter_conditions:
                filtered_df = df[filter_conditions(df)]
            else:
                filtered_df = df
            
            # 返回前三列
            result = filtered_df.iloc[:, [0, 1, 2]]
            result_list = result.values.tolist()
            return result_list
            
        except Exception as e:
            print(f"读取过滤数据错误: {e}")
            return []

    def highlight_core_stocks(self, source_sheet_name, target_sheet_name):
        """
        将源工作表中核心列等于1的股票名称在目标工作表中设为红色
        
        Args:
            source_sheet_name: 源工作表名称（包含核心列）
            target_sheet_name: 目标工作表名称（watch工作表）
        """
        try:
            # 读取源工作表数据
            source_df = self.read_data_from_sheet(source_sheet_name)
            if source_df is None:
                print(f"无法读取源工作表: {source_sheet_name}")
                return
            
            # 检查是否有核心列
            if '核心' not in source_df.columns:
                print(f"源工作表 {source_sheet_name} 中未找到'核心'列")
                return
            
            # 获取核心列等于1的股票名称
            core_stocks = source_df[source_df['核心'] == 1]['名称'].tolist()
            
            if not core_stocks:
                print(f"源工作表 {source_sheet_name} 中没有核心列等于1的股票")
                return
            
            print(f"找到核心股票: {core_stocks}")
            
            # 打开Excel文件
            wb = load_workbook(self.file_path)
            ws = wb[target_sheet_name]
            
            # 获取名称列的索引
            header_row = list(ws.iter_rows(max_row=1))[0]
            name_col_index = None
            for idx, cell in enumerate(header_row):
                if cell.value == '名称':
                    name_col_index = idx
                    break
            
            if name_col_index is None:
                print(f"目标工作表 {target_sheet_name} 中未找到'名称'列")
                wb.close()
                return
            
            # 遍历目标工作表的名称列，将核心股票设为红色
            red_font = Font(color="FF0000", bold=True)  # 红色加粗字体
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                name_cell = row[name_col_index]
                if name_cell.value in core_stocks:
                    name_cell.font = red_font
                    print(f"已将股票 {name_cell.value} 设为红色")
            
            # 保存文件
            wb.save(self.file_path)
            wb.close()
            print(f"核心股票高亮完成，共处理 {len(core_stocks)} 只股票")
            
        except Exception as e:
            print(f"高亮核心股票错误: {e}")